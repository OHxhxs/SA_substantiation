"""
엑셀 파일의 시트들을 읽어서 JSON으로 변환하는 스크립트
openpyxl을 직접 사용
"""
from openpyxl import load_workbook
import json
from pathlib import Path

def convert_excel_to_json():
    excel_file = 'interview_system_db.xlsx'
    output_dir = Path('frontend/src/data')
    output_dir.mkdir(parents=True, exist_ok=True)
    
    # 엑셀 파일 읽기
    wb = load_workbook(excel_file, data_only=True)
    
    print(f"발견된 시트: {wb.sheetnames}")
    
    data = {}
    
    # 각 시트를 읽어서 JSON으로 변환
    for sheet_name in wb.sheetnames:
        print(f"\n처리 중: {sheet_name}")
        ws = wb[sheet_name]
        
        # 첫 번째 행을 헤더로 사용
        headers = []
        for cell in ws[1]:
            headers.append(cell.value)
        
        print(f"  - 컬럼: {headers}")
        
        # 데이터 행들을 딕셔너리 리스트로 변환
        sheet_data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            # 빈 행 건너뛰기
            if all(cell is None for cell in row):
                continue
            
            row_dict = {}
            for header, value in zip(headers, row):
                if header:  # 헤더가 None이 아닌 경우만
                    row_dict[header] = value
            
            sheet_data.append(row_dict)
        
        print(f"  - {len(sheet_data)}개 레코드 발견")
        
        data[sheet_name] = sheet_data
    
    # JSON 파일로 저장
    output_file = output_dir / 'questionnaireData.json'
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    
    print(f"\n✅ 변환 완료: {output_file}")
    print(f"총 {len(data)} 개의 시트 처리됨")
    
    return data

if __name__ == "__main__":
    data = convert_excel_to_json()
    
    # 데이터 구조 출력
    print("\n" + "="*50)
    print("데이터 구조 요약:")
    print("="*50)
    for sheet_name, records in data.items():
        print(f"\n[{sheet_name}]: {len(records)} records")
        if len(records) > 0:
            print(f"  샘플 데이터 (첫 번째): {records[0]}")
